#!/usr/bin/env python3
"""
Generate a test shipment document (PDF or Excel) for testing document scraping
"""

import os
from datetime import datetime, timedelta

# Try to use reportlab for PDF generation
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfgen import canvas
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Try to use pandas for Excel generation
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

# Try to use fpdf for PDF generation (simpler alternative)
try:
    from fpdf import FPDF
    FPDF_AVAILABLE = True
except ImportError:
    FPDF_AVAILABLE = False

def generate_test_shipment_pdf(output_path='test_shipment.pdf'):
    """
    Generate a test shipment PDF with sample data
    """
    if not REPORTLAB_AVAILABLE:
        # Fallback: Create a simple text-based approach or use another library
        print("reportlab is required for PDF generation.")
        print("Please install it with: pip install reportlab")
        return None
    
    # Sample shipment data
    sample_items = [
        {
            'sku': 'SKU-001',
            'product_name': 'Organic Apples - Red Delicious',
            'quantity': 24,
            'unit_price': 2.99,
            'lot_number': 'LOT-2024-001',
            'expiration_date': '2024-12-31'
        },
        {
            'sku': 'SKU-002',
            'product_name': 'Fresh Bananas - Organic',
            'quantity': 36,
            'unit_price': 1.49,
            'lot_number': 'LOT-2024-002',
            'expiration_date': '2024-11-15'
        },
        {
            'sku': 'SKU-003',
            'product_name': 'Carrots - Baby Cut',
            'quantity': 48,
            'unit_price': 3.99,
            'lot_number': 'LOT-2024-003',
            'expiration_date': '2024-12-20'
        },
        {
            'sku': 'SKU-004',
            'product_name': 'Lettuce - Romaine Hearts',
            'quantity': 12,
            'unit_price': 4.99,
            'lot_number': 'LOT-2024-004',
            'expiration_date': '2024-11-10'
        },
        {
            'sku': 'SKU-005',
            'product_name': 'Tomatoes - Cherry',
            'quantity': 30,
            'unit_price': 5.99,
            'lot_number': 'LOT-2024-005',
            'expiration_date': '2024-11-25'
        },
        {
            'sku': 'SKU-006',
            'product_name': 'Broccoli - Fresh Crowns',
            'quantity': 20,
            'unit_price': 3.49,
            'lot_number': 'LOT-2024-006',
            'expiration_date': '2024-12-05'
        },
        {
            'sku': 'SKU-007',
            'product_name': 'Spinach - Baby Leaf',
            'quantity': 18,
            'unit_price': 4.49,
            'lot_number': 'LOT-2024-007',
            'expiration_date': '2024-11-18'
        },
        {
            'sku': 'SKU-008',
            'product_name': 'Bell Peppers - Mixed Colors',
            'quantity': 15,
            'unit_price': 6.99,
            'lot_number': 'LOT-2024-008',
            'expiration_date': '2024-12-10'
        },
        {
            'sku': 'SKU-009',
            'product_name': 'Cucumbers - English',
            'quantity': 25,
            'unit_price': 2.99,
            'lot_number': 'LOT-2024-009',
            'expiration_date': '2024-11-30'
        },
        {
            'sku': 'SKU-010',
            'product_name': 'Onions - Yellow',
            'quantity': 40,
            'unit_price': 1.99,
            'lot_number': 'LOT-2024-010',
            'expiration_date': '2025-01-15'
        }
    ]
    
    # Create PDF document
    doc = SimpleDocTemplate(output_path, pagesize=letter)
    story = []
    
    # Define styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=30,
        alignment=1  # Center
    )
    
    # Add title
    story.append(Paragraph("SHIPMENT SUMMARY", title_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Add header information
    header_data = [
        ['Vendor:', 'Fresh Produce Co.'],
        ['Purchase Order:', 'PO-2024-001'],
        ['Shipment Date:', datetime.now().strftime('%Y-%m-%d')],
        ['Expected Delivery:', (datetime.now() + timedelta(days=3)).strftime('%Y-%m-%d')]
    ]
    
    header_table = Table(header_data, colWidths=[2*inch, 4*inch])
    header_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Prepare table data
    table_data = [['SKU', 'Product Name', 'Quantity', 'Unit Price', 'Lot Number', 'Expiration Date']]
    
    total_quantity = 0
    total_cost = 0.0
    
    for item in sample_items:
        row = [
            item['sku'],
            item['product_name'],
            str(item['quantity']),
            f"${item['unit_price']:.2f}",
            item['lot_number'],
            item['expiration_date']
        ]
        table_data.append(row)
        total_quantity += item['quantity']
        total_cost += item['quantity'] * item['unit_price']
    
    # Add totals row
    table_data.append([
        'TOTAL',
        '',
        str(total_quantity),
        f"${total_cost:.2f}",
        '',
        ''
    ])
    
    # Create table
    table = Table(table_data, colWidths=[1*inch, 3*inch, 0.8*inch, 1*inch, 1.2*inch, 1.2*inch])
    
    # Style the table
    table.setStyle(TableStyle([
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (2, 1), (3, -2), 'RIGHT'),  # Numbers right-aligned
        ('ALIGN', (2, -1), (3, -1), 'RIGHT'),  # Total row
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        
        # Data rows
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F2F2F2')]),
        
        # Total row
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D9E1F2')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 10),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#2F5597')),
        ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#2F5597')),
        
        # Padding
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))
    
    story.append(table)
    story.append(Spacer(1, 0.3*inch))
    
    # Add summary
    summary_text = f"""
    <b>Summary:</b><br/>
    Total Items: {len(sample_items)}<br/>
    Total Quantity: {total_quantity} units<br/>
    Total Cost: ${total_cost:.2f}
    """
    story.append(Paragraph(summary_text, styles['Normal']))
    
    # Build PDF
    doc.build(story)
    
    return output_path

def generate_test_shipment_pdf_fpdf(output_path='test_shipment.pdf'):
    """
    Generate a test shipment PDF using fpdf (simpler library)
    """
    if not FPDF_AVAILABLE:
        return None
    
    # Sample shipment data
    sample_items = [
        {'sku': 'SKU-001', 'product_name': 'Organic Apples - Red Delicious', 'quantity': 24, 'unit_price': 2.99, 'lot_number': 'LOT-2024-001', 'expiration_date': '2024-12-31'},
        {'sku': 'SKU-002', 'product_name': 'Fresh Bananas - Organic', 'quantity': 36, 'unit_price': 1.49, 'lot_number': 'LOT-2024-002', 'expiration_date': '2024-11-15'},
        {'sku': 'SKU-003', 'product_name': 'Carrots - Baby Cut', 'quantity': 48, 'unit_price': 3.99, 'lot_number': 'LOT-2024-003', 'expiration_date': '2024-12-20'},
        {'sku': 'SKU-004', 'product_name': 'Lettuce - Romaine Hearts', 'quantity': 12, 'unit_price': 4.99, 'lot_number': 'LOT-2024-004', 'expiration_date': '2024-11-10'},
        {'sku': 'SKU-005', 'product_name': 'Tomatoes - Cherry', 'quantity': 30, 'unit_price': 5.99, 'lot_number': 'LOT-2024-005', 'expiration_date': '2024-11-25'},
        {'sku': 'SKU-006', 'product_name': 'Broccoli - Fresh Crowns', 'quantity': 20, 'unit_price': 3.49, 'lot_number': 'LOT-2024-006', 'expiration_date': '2024-12-05'},
        {'sku': 'SKU-007', 'product_name': 'Spinach - Baby Leaf', 'quantity': 18, 'unit_price': 4.49, 'lot_number': 'LOT-2024-007', 'expiration_date': '2024-11-18'},
        {'sku': 'SKU-008', 'product_name': 'Bell Peppers - Mixed Colors', 'quantity': 15, 'unit_price': 6.99, 'lot_number': 'LOT-2024-008', 'expiration_date': '2024-12-10'},
        {'sku': 'SKU-009', 'product_name': 'Cucumbers - English', 'quantity': 25, 'unit_price': 2.99, 'lot_number': 'LOT-2024-009', 'expiration_date': '2024-11-30'},
        {'sku': 'SKU-010', 'product_name': 'Onions - Yellow', 'quantity': 40, 'unit_price': 1.99, 'lot_number': 'LOT-2024-010', 'expiration_date': '2025-01-15'},
    ]
    
    pdf = FPDF()
    pdf.add_page()
    
    # Title
    pdf.set_font('Arial', 'B', 20)
    pdf.cell(0, 10, 'SHIPMENT SUMMARY', 0, 1, 'C')
    pdf.ln(5)
    
    # Header info
    pdf.set_font('Arial', '', 10)
    pdf.cell(0, 5, f'Vendor: Fresh Produce Co.', 0, 1)
    pdf.cell(0, 5, f'Purchase Order: PO-2024-001', 0, 1)
    pdf.cell(0, 5, f'Shipment Date: {datetime.now().strftime("%Y-%m-%d")}', 0, 1)
    pdf.cell(0, 5, f'Expected Delivery: {(datetime.now() + timedelta(days=3)).strftime("%Y-%m-%d")}', 0, 1)
    pdf.ln(5)
    
    # Table header
    pdf.set_font('Arial', 'B', 10)
    col_widths = [30, 80, 25, 30, 35, 40]
    headers = ['SKU', 'Product Name', 'Qty', 'Price', 'Lot #', 'Exp Date']
    
    for i, header in enumerate(headers):
        pdf.cell(col_widths[i], 8, header, 1, 0, 'C')
    pdf.ln()
    
    # Table data
    pdf.set_font('Arial', '', 9)
    total_qty = 0
    total_cost = 0.0
    
    for item in sample_items:
        pdf.cell(col_widths[0], 6, item['sku'], 1, 0, 'L')
        pdf.cell(col_widths[1], 6, item['product_name'][:35], 1, 0, 'L')  # Truncate long names
        pdf.cell(col_widths[2], 6, str(item['quantity']), 1, 0, 'R')
        pdf.cell(col_widths[3], 6, f"${item['unit_price']:.2f}", 1, 0, 'R')
        pdf.cell(col_widths[4], 6, item['lot_number'], 1, 0, 'L')
        pdf.cell(col_widths[5], 6, item['expiration_date'], 1, 0, 'L')
        pdf.ln()
        
        total_qty += item['quantity']
        total_cost += item['quantity'] * item['unit_price']
    
    # Total row
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(col_widths[0] + col_widths[1], 6, 'TOTAL', 1, 0, 'R')
    pdf.cell(col_widths[2], 6, str(total_qty), 1, 0, 'R')
    pdf.cell(col_widths[3], 6, f"${total_cost:.2f}", 1, 0, 'R')
    pdf.cell(col_widths[4] + col_widths[5], 6, '', 1, 0, 'L')
    pdf.ln(10)
    
    # Summary
    pdf.set_font('Arial', '', 10)
    pdf.cell(0, 5, f'Summary: {len(sample_items)} items, {total_qty} total units, ${total_cost:.2f} total cost', 0, 1)
    
    pdf.output(output_path)
    return output_path

def generate_simple_pdf_alternative(output_path='test_shipment.pdf'):
    """
    Alternative method using basic text file that can be converted to PDF
    or using a simpler approach
    """
    sample_items = [
        {'sku': 'SKU-001', 'product_name': 'Organic Apples', 'quantity': 24, 'unit_price': 2.99},
        {'sku': 'SKU-002', 'product_name': 'Fresh Bananas', 'quantity': 36, 'unit_price': 1.49},
        {'sku': 'SKU-003', 'product_name': 'Carrots', 'quantity': 48, 'unit_price': 3.99},
    ]
    
    # Create a simple text file (user can convert to PDF manually)
    txt_path = output_path.replace('.pdf', '.txt')
    with open(txt_path, 'w') as f:
        f.write("SHIPMENT SUMMARY\n")
        f.write("=" * 50 + "\n\n")
        f.write(f"Vendor: Fresh Produce Co.\n")
        f.write(f"Purchase Order: PO-2024-001\n")
        f.write(f"Date: {datetime.now().strftime('%Y-%m-%d')}\n\n")
        f.write("SKU\tProduct Name\tQuantity\tUnit Price\n")
        f.write("-" * 50 + "\n")
        for item in sample_items:
            f.write(f"{item['sku']}\t{item['product_name']}\t{item['quantity']}\t${item['unit_price']:.2f}\n")
    
    print(f"Created text file: {txt_path}")
    print("Note: Install reportlab for PDF generation: pip install reportlab")
    return txt_path

def generate_test_shipment_excel(output_path='test_shipment.xlsx'):
    """
    Generate a test shipment Excel file (alternative to PDF)
    """
    if not PANDAS_AVAILABLE:
        print("pandas not available. Install with: pip install pandas openpyxl")
        return None
    
    # Sample shipment data
    sample_items = [
        {
            'SKU': 'SKU-001',
            'Product Name': 'Organic Apples - Red Delicious',
            'Quantity': 24,
            'Unit Price': 2.99,
            'Lot Number': 'LOT-2024-001',
            'Expiration Date': '2024-12-31'
        },
        {
            'SKU': 'SKU-002',
            'Product Name': 'Fresh Bananas - Organic',
            'Quantity': 36,
            'Unit Price': 1.49,
            'Lot Number': 'LOT-2024-002',
            'Expiration Date': '2024-11-15'
        },
        {
            'SKU': 'SKU-003',
            'Product Name': 'Carrots - Baby Cut',
            'Quantity': 48,
            'Unit Price': 3.99,
            'Lot Number': 'LOT-2024-003',
            'Expiration Date': '2024-12-20'
        },
        {
            'SKU': 'SKU-004',
            'Product Name': 'Lettuce - Romaine Hearts',
            'Quantity': 12,
            'Unit Price': 4.99,
            'Lot Number': 'LOT-2024-004',
            'Expiration Date': '2024-11-10'
        },
        {
            'SKU': 'SKU-005',
            'Product Name': 'Tomatoes - Cherry',
            'Quantity': 30,
            'Unit Price': 5.99,
            'Lot Number': 'LOT-2024-005',
            'Expiration Date': '2024-11-25'
        },
        {
            'SKU': 'SKU-006',
            'Product Name': 'Broccoli - Fresh Crowns',
            'Quantity': 20,
            'Unit Price': 3.49,
            'Lot Number': 'LOT-2024-006',
            'Expiration Date': '2024-12-05'
        },
        {
            'SKU': 'SKU-007',
            'Product Name': 'Spinach - Baby Leaf',
            'Quantity': 18,
            'Unit Price': 4.49,
            'Lot Number': 'LOT-2024-007',
            'Expiration Date': '2024-11-18'
        },
        {
            'SKU': 'SKU-008',
            'Product Name': 'Bell Peppers - Mixed Colors',
            'Quantity': 15,
            'Unit Price': 6.99,
            'Lot Number': 'LOT-2024-008',
            'Expiration Date': '2024-12-10'
        },
        {
            'SKU': 'SKU-009',
            'Product Name': 'Cucumbers - English',
            'Quantity': 25,
            'Unit Price': 2.99,
            'Lot Number': 'LOT-2024-009',
            'Expiration Date': '2024-11-30'
        },
        {
            'SKU': 'SKU-010',
            'Product Name': 'Onions - Yellow',
            'Quantity': 40,
            'Unit Price': 1.99,
            'Lot Number': 'LOT-2024-010',
            'Expiration Date': '2025-01-15'
        }
    ]
    
    # Create DataFrame
    df = pd.DataFrame(sample_items)
    
    # Create Excel writer with formatting
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Shipment', index=False)
        
        # Get the worksheet
        worksheet = writer.sheets['Shipment']
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Style header row
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add totals row
        total_row = len(df) + 3
        worksheet.cell(row=total_row, column=1, value='TOTAL')
        worksheet.cell(row=total_row, column=3, value=df['Quantity'].sum())
        worksheet.cell(row=total_row, column=4, value=(df['Quantity'] * df['Unit Price']).sum())
        
        total_cell = worksheet.cell(row=total_row, column=1)
        total_cell.font = Font(bold=True)
        worksheet.cell(row=total_row, column=4).font = Font(bold=True)
    
    return output_path

if __name__ == '__main__':
    import sys
    
    output_file = sys.argv[1] if len(sys.argv) > 1 else 'test_shipment.pdf'
    
    # Determine output format
    if output_file.endswith('.xlsx') or output_file.endswith('.xls'):
        if PANDAS_AVAILABLE:
            result = generate_test_shipment_excel(output_file)
            if result:
                print(f"✓ Successfully generated test shipment Excel: {result}")
                print(f"  File size: {os.path.getsize(result)} bytes")
                print(f"  You can use this file to test the document upload feature!")
        else:
            print("pandas/openpyxl not available. Install with: pip install pandas openpyxl")
    elif output_file.endswith('.pdf'):
        if REPORTLAB_AVAILABLE:
            result = generate_test_shipment_pdf(output_file)
            if result:
                print(f"✓ Successfully generated test shipment PDF: {result}")
                print(f"  File size: {os.path.getsize(result)} bytes")
        elif FPDF_AVAILABLE:
            result = generate_test_shipment_pdf_fpdf(output_file)
            if result:
                print(f"✓ Successfully generated test shipment PDF: {result}")
                print(f"  File size: {os.path.getsize(result)} bytes")
        else:
            print("PDF libraries not available.")
            print("Generating Excel file instead (can be converted to PDF)...")
            excel_file = output_file.replace('.pdf', '.xlsx')
            if PANDAS_AVAILABLE:
                result = generate_test_shipment_excel(excel_file)
                if result:
                    print(f"✓ Created Excel file: {result}")
                    print(f"  You can convert this to PDF using Excel or use it directly for testing.")
            else:
                print("No PDF/Excel libraries available.")
                print("Install one of:")
                print("  pip install reportlab  (for PDF)")
                print("  pip install fpdf2  (for PDF - simpler)")
                print("  pip install pandas openpyxl  (for Excel)")
    else:
        # Default: try PDF first, then Excel
        if REPORTLAB_AVAILABLE:
            result = generate_test_shipment_pdf(output_file)
            if result:
                print(f"✓ Successfully generated test shipment PDF: {result}")
        elif PANDAS_AVAILABLE:
            excel_file = output_file.replace('.pdf', '.xlsx') if output_file.endswith('.pdf') else output_file + '.xlsx'
            result = generate_test_shipment_excel(excel_file)
            if result:
                print(f"✓ Created Excel file: {result}")
        else:
            print("No PDF/Excel libraries available.")
            print("Install one of:")
            print("  pip install reportlab  (for PDF)")
            print("  pip install pandas openpyxl  (for Excel)")

